import openpyxl

#como criar ujma planilha(book)
book = openpyxl.Workbook()
#como visualizar páginas existentes
print(book.sheetnames)
#como criar uma página
book.create_sheet('Frutas')

# como selecionar uma página
frutas_page = book['Frutas']
frutas_page.append(['FRUTA', 'QUANTIDADE', 'PREÇO'])
frutas_page.append(['banana', '5', 'R$3,00'])
frutas_page.append(['maçã', '15', 'R$1,20'])
frutas_page.append(['abacaxi', '7', 'R$5,00'])
frutas_page.append(['uva', '35', 'R$1,00'])

# Salvar a planilha
book.save('planilhadecompras.xlsx')
